# -*- coding: utf-8 -*-
"""
Excel Service
- Excel dosyası okuma
- Çalışan verisi parse etme
- Rapor oluşturma
"""

from typing import List, Dict, Tuple, Optional
from io import BytesIO
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False


class ExcelService:
    """Excel işlemlerini yöneten servis"""
    
    # Tanınan sütun isimleri
    TC_HEADERS = ['tc', 'tckn', 'tc kimlik', 'tc no', 'tckimlik', 't.c. kimlik no', 'tc kimlik no', 'kimlik no']
    MAIL_HEADERS = ['mail', 'email', 'e-mail', 'eposta', 'e-posta', 'mail adresi', 'e-posta adresi']
    NAME_HEADERS = ['personel adı soyadı', 'ad soyad', 'adı soyadı', 'isim', 'ad', 'personel', 'çalışan', 'isim soyisim']
    FIRST_NAME_HEADERS = ['ad', 'isim', 'first name', 'firstname']
    LAST_NAME_HEADERS = ['soyad', 'soyadı', 'last name', 'lastname']
    DEPARTMENT_HEADERS = ['departman', 'bölüm', 'department', 'birim']
    
    def __init__(self):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl modülü yüklü değil")
    
    def read_employees_from_excel(
        self, 
        content: bytes, 
        file_ext: str
    ) -> Tuple[List[Dict], List[str]]:
        """
        Excel dosyasından çalışan verilerini oku
        
        Args:
            content: Dosya içeriği (bytes)
            file_ext: Dosya uzantısı (xlsx veya xls)
            
        Returns:
            Tuple[List[Dict], List[str]]: (çalışan listesi, hatalar)
        """
        if file_ext == "xls":
            return self._read_xls(content)
        else:
            return self._read_xlsx(content)
    
    def _find_column(self, headers: List, search_list: List[str]) -> Optional[int]:
        """Başlıklar içinde eşleşen sütunu bul"""
        for idx, header in enumerate(headers):
            if header:
                header_lower = str(header).lower().strip()
                if header_lower in search_list:
                    return idx
        return None
    
    def _read_xlsx(self, content: bytes) -> Tuple[List[Dict], List[str]]:
        """xlsx formatını oku"""
        employees = []
        errors = []
        
        try:
            wb = openpyxl.load_workbook(BytesIO(content), read_only=True)
            ws = wb.active
            
            # Başlık satırını kontrol et
            headers = [cell.value for cell in ws[1]]
            
            # Sütunları bul
            tc_col = self._find_column(headers, self.TC_HEADERS)
            mail_col = self._find_column(headers, self.MAIL_HEADERS)
            name_col = self._find_column(headers, self.NAME_HEADERS)
            first_name_col = self._find_column(headers, self.FIRST_NAME_HEADERS)
            last_name_col = self._find_column(headers, self.LAST_NAME_HEADERS)
            department_col = self._find_column(headers, self.DEPARTMENT_HEADERS)
            
            if tc_col is None:
                errors.append("Excel'de TC sütunu bulunamadı")
                return [], errors
            
            if mail_col is None:
                errors.append("Excel'de Mail sütunu bulunamadı")
                return [], errors
            
            # Verileri oku
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                result = self._process_row(
                    row, row_idx, tc_col, mail_col, name_col,
                    first_name_col, last_name_col, department_col, errors
                )
                if result:
                    employees.append(result)
            
            wb.close()
            
        except Exception as e:
            errors.append(f"Excel okuma hatası: {str(e)}")
        
        return employees, errors
    
    def _read_xls(self, content: bytes) -> Tuple[List[Dict], List[str]]:
        """xls formatını oku"""
        employees = []
        errors = []
        
        if not XLRD_AVAILABLE:
            errors.append("xls formatı için xlrd modülü gerekli")
            return [], errors
        
        try:
            wb = xlrd.open_workbook(file_contents=content)
            ws = wb.sheet_by_index(0)
            
            # Başlık satırını oku
            headers = [ws.cell_value(0, col) for col in range(ws.ncols)]
            
            # Sütunları bul
            tc_col = self._find_column(headers, self.TC_HEADERS)
            mail_col = self._find_column(headers, self.MAIL_HEADERS)
            name_col = self._find_column(headers, self.NAME_HEADERS)
            first_name_col = self._find_column(headers, self.FIRST_NAME_HEADERS)
            last_name_col = self._find_column(headers, self.LAST_NAME_HEADERS)
            department_col = self._find_column(headers, self.DEPARTMENT_HEADERS)
            
            if tc_col is None:
                errors.append("Excel'de TC sütunu bulunamadı")
                return [], errors
            
            if mail_col is None:
                errors.append("Excel'de Mail sütunu bulunamadı")
                return [], errors
            
            # Verileri oku
            for row_idx in range(1, ws.nrows):
                row = [ws.cell_value(row_idx, col) for col in range(ws.ncols)]
                result = self._process_row(
                    row, row_idx + 1, tc_col, mail_col, name_col,
                    first_name_col, last_name_col, department_col, errors
                )
                if result:
                    employees.append(result)
            
        except Exception as e:
            errors.append(f"Excel okuma hatası: {str(e)}")
        
        return employees, errors
    
    def _process_row(
        self, 
        row, 
        row_idx, 
        tc_col, 
        mail_col, 
        name_col,
        first_name_col,
        last_name_col,
        department_col,
        errors
    ) -> Optional[Dict]:
        """Satır verisini işle"""
        # TC'yi al
        tc_value = row[tc_col] if tc_col < len(row) else None
        mail_value = row[mail_col] if mail_col < len(row) else None
        
        # Boş satırları atla
        if not tc_value and not mail_value:
            return None
        
        # TC'yi string'e çevir ve temizle
        tc_str = str(tc_value).strip() if tc_value else ""
        
        # TC float olarak gelebilir (12345678901.0)
        if '.' in tc_str:
            tc_str = tc_str.split('.')[0]
        
        # TC sadece rakamlardan oluşmalı
        tc_clean = ''.join(filter(str.isdigit, tc_str))
        
        # Mail'i temizle
        mail_str = str(mail_value).strip() if mail_value else ""
        
        # İsimleri al
        first_name = None
        last_name = None
        
        if first_name_col is not None and first_name_col < len(row):
            first_name = str(row[first_name_col]).strip() if row[first_name_col] else None
        
        if last_name_col is not None and last_name_col < len(row):
            last_name = str(row[last_name_col]).strip() if row[last_name_col] else None
        
        # Ad soyad birlikte ise ayır
        if name_col is not None and name_col < len(row) and row[name_col]:
            full_name = str(row[name_col]).strip()
            if full_name and not first_name and not last_name:
                parts = full_name.split()
                if len(parts) >= 2:
                    first_name = " ".join(parts[:-1])
                    last_name = parts[-1]
                elif len(parts) == 1:
                    first_name = parts[0]
        
        # Departman
        department = None
        if department_col is not None and department_col < len(row):
            department = str(row[department_col]).strip() if row[department_col] else None
        
        # Doğrulama
        if not tc_clean:
            errors.append(f"Satır {row_idx}: TC boş veya geçersiz")
            return None
        
        if len(tc_clean) != 11:
            errors.append(f"Satır {row_idx}: TC 11 haneli olmalı")
            return None
        
        if not mail_str or '@' not in mail_str:
            errors.append(f"Satır {row_idx}: Mail adresi geçersiz")
            return None
        
        # İsim zorunlu kontrolü
        if not first_name and not last_name:
            errors.append(f"Satır {row_idx}: Personel Adı Soyadı boş olamaz")
            return None
        
        return {
            "tc_no": tc_clean,
            "email": mail_str,
            "first_name": first_name,
            "last_name": last_name,
            "department": department
        }
    
    def create_send_report(self, results: List[Dict]) -> bytes:
        """
        Mail gönderim raporunu Excel olarak oluştur
        
        Args:
            results: Gönderim sonuçları listesi
                Her sonuç: {
                    "employee_name": str,
                    "employee_email": str,
                    "tc_no": str,
                    "period": str,
                    "status": str,  # "Başarılı", "Hatalı", "Çalışan Yok"
                    "error": str veya None,
                    "sent_at": datetime veya None,
                    "opened_at": datetime veya None,
                    "downloaded_at": datetime veya None,
                    "download_count": int
                }
                
        Returns:
            bytes: Excel dosyası içeriği
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Gönderim Raporu"
        
        # Stiller
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        
        # Durum renkleri
        downloaded_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")  # Yeşil - İndirildi
        opened_fill = PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid")  # Mavi - Okundu
        sent_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")  # Sarı - Gönderildi
        no_employee_fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")  # Turuncu - Çalışan Yok
        failed_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")  # Kırmızı - Başarısız
        pending_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")  # Gri - Bekliyor
        
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Başlıklar
        headers = ["Sıra", "Ad Soyad", "TC (Son 4)", "E-Posta", "Dönem", "Durum", "Gönderim", "Okunma", "İndirme", "İndirme Sayısı", "Hata Açıklaması"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        # Veriler
        for row_idx, result in enumerate(results, start=2):
            # Sıra
            ws.cell(row=row_idx, column=1, value=row_idx - 1).alignment = center_align
            
            # Ad Soyad
            ws.cell(row=row_idx, column=2, value=result.get("employee_name", "-")).alignment = left_align
            
            # TC (maskelenmiş)
            tc = result.get("tc_no", "")
            tc_masked = f"****{tc[-4:]}" if tc and len(tc) >= 4 else "-"
            ws.cell(row=row_idx, column=3, value=tc_masked).alignment = center_align
            
            # E-Posta
            ws.cell(row=row_idx, column=4, value=result.get("employee_email", "-")).alignment = left_align
            
            # Dönem
            ws.cell(row=row_idx, column=5, value=result.get("period", "-")).alignment = center_align
            
            # Durum
            status = result.get("status", "Bilinmiyor")
            status_cell = ws.cell(row=row_idx, column=6, value=status)
            status_cell.alignment = center_align
            
            # Durum rengini ayarla
            if status == "İndirildi":
                row_fill = downloaded_fill
            elif status == "Okundu":
                row_fill = opened_fill
            elif status == "Gönderildi":
                row_fill = sent_fill
            elif status == "Çalışan Yok":
                row_fill = no_employee_fill
            elif status == "Başarısız":
                row_fill = failed_fill
            else:
                row_fill = pending_fill
            
            status_cell.fill = row_fill
            
            # Gönderim zamanı
            sent_at = result.get("sent_at")
            sent_str = sent_at.strftime("%d.%m.%Y %H:%M") if sent_at else "-"
            ws.cell(row=row_idx, column=7, value=sent_str).alignment = center_align
            
            # Okunma zamanı
            opened_at = result.get("opened_at")
            opened_str = opened_at.strftime("%d.%m.%Y %H:%M") if opened_at else "-"
            ws.cell(row=row_idx, column=8, value=opened_str).alignment = center_align
            
            # İndirme zamanı
            downloaded_at = result.get("downloaded_at")
            downloaded_str = downloaded_at.strftime("%d.%m.%Y %H:%M") if downloaded_at else "-"
            ws.cell(row=row_idx, column=9, value=downloaded_str).alignment = center_align
            
            # İndirme sayısı
            download_count = result.get("download_count", 0)
            ws.cell(row=row_idx, column=10, value=download_count).alignment = center_align
            
            # Hata açıklaması
            ws.cell(row=row_idx, column=11, value=result.get("error") or "-").alignment = left_align
            
            # Border uygula
            for col in range(1, 12):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = thin_border
        
        # Sütun genişlikleri
        column_widths = [8, 25, 12, 35, 15, 15, 18, 18, 18, 15, 40]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        
        # Özet bilgiler
        summary_row = len(results) + 3
        
        total = len(results)
        downloaded_count = sum(1 for r in results if r.get("status") == "İndirildi")
        opened_count = sum(1 for r in results if r.get("status") == "Okundu")
        sent_count = sum(1 for r in results if r.get("status") == "Gönderildi")
        no_employee = sum(1 for r in results if r.get("status") == "Çalışan Yok")
        failed = sum(1 for r in results if r.get("status") == "Başarısız")
        pending = sum(1 for r in results if r.get("status") == "Bekliyor")
        
        # Toplam okunan ve indirilen (her durumda)
        total_opened = sum(1 for r in results if r.get("opened_at"))
        total_downloaded = sum(1 for r in results if r.get("downloaded_at"))
        
        ws.cell(row=summary_row, column=1, value="ÖZET").font = Font(bold=True, size=12)
        ws.cell(row=summary_row + 1, column=1, value=f"Toplam Kayıt: {total}").font = Font(bold=True)
        ws.cell(row=summary_row + 2, column=1, value="")
        
        ws.cell(row=summary_row + 3, column=1, value="DURUM DAĞILIMI:").font = Font(bold=True)
        ws.cell(row=summary_row + 4, column=1, value=f"İndirildi: {downloaded_count}").fill = downloaded_fill
        ws.cell(row=summary_row + 5, column=1, value=f"Okundu: {opened_count}").fill = opened_fill
        ws.cell(row=summary_row + 6, column=1, value=f"Gönderildi: {sent_count}").fill = sent_fill
        ws.cell(row=summary_row + 7, column=1, value=f"Çalışan Yok: {no_employee}").fill = no_employee_fill
        ws.cell(row=summary_row + 8, column=1, value=f"Başarısız: {failed}").fill = failed_fill
        ws.cell(row=summary_row + 9, column=1, value=f"Bekliyor: {pending}").fill = pending_fill
        
        ws.cell(row=summary_row + 11, column=1, value="İSTATİSTİKLER:").font = Font(bold=True)
        
        # Açılma ve indirme oranları hesapla
        sent_total = total - no_employee - pending  # Gönderilen toplam
        open_rate = round((total_opened / sent_total * 100), 1) if sent_total > 0 else 0
        download_rate = round((total_downloaded / sent_total * 100), 1) if sent_total > 0 else 0
        
        ws.cell(row=summary_row + 12, column=1, value=f"Gönderilen Toplam: {sent_total}")
        ws.cell(row=summary_row + 13, column=1, value=f"Açılan Toplam: {total_opened} (%{open_rate})")
        ws.cell(row=summary_row + 14, column=1, value=f"İndirilen Toplam: {total_downloaded} (%{download_rate})")
        
        # Rapor oluşturma tarihi
        ws.cell(row=summary_row + 16, column=1, value=f"Rapor Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
        
        # Sütun A genişliğini özet için ayarla
        ws.column_dimensions['A'].width = max(column_widths[0], 35)
        
        # BytesIO'ya kaydet
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()



